from datetime import date

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def absentee_highlighter(input_filename):

    # Read in input file
    full_membership_workbook = load_workbook(input_filename)
    full_membership_worksheet = full_membership_workbook.active
    row_count = 2  # Skip first row, which contains headers

    # Set up destination file
    absentee_highlighted_workbook = openpyxl.Workbook()
    absentee_highlighted_worksheet = absentee_highlighted_workbook.active

    # Set up header rows on destination file
    columns = ["A", "B", "C"]
    header_names = ["Full Name", "Phone", "Days Absent"]

    for index in range(3):
        cell = absentee_highlighted_worksheet.cell(1, index + 1)

        cell.value = header_names[index]
        cell.border = Border(bottom=Side(style="medium"))
        cell.font = Font(bold=True)

        absentee_highlighted_worksheet.column_dimensions[columns[index]].width = 20

    # Iterate through membership worksheet and add rows to destination worksheet
    while True:
        # Get values from columns
        full_name = full_membership_worksheet.cell(row_count, 1).value
        phone = full_membership_worksheet.cell(row_count, 4).value
        days_absent = full_membership_worksheet.cell(row_count, 8).value

        # Check End of File
        if full_name is None:
            break

        # Copy values into highlighted worksheet
        absentee_highlighted_worksheet.cell(row_count, 1).value = full_name
        absentee_highlighted_worksheet.cell(row_count, 2).value = phone
        absentee_highlighted_worksheet.cell(row_count, 3).value = days_absent

        # If `days_absent` is excessive, highlight the cell
        if days_absent > 10:
            absentee_highlighted_worksheet.cell(row_count, 3).fill = PatternFill("solid", fgColor="0099CC00")

        row_count += 1

    # Save file!
    absentee_highlighted_workbook.save(filename="excessive-absentees-highlighted-" +
                                                date.today().strftime("%m%d%y") +
                                                ".xlsx")


if __name__ == '__main__':
    absentee_highlighter('full-membership.xlsx')
